import io
import tempfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.sql import Select

from app.deps import CurrentUser, DbSession, project_access
from app.extractors.topside_excel import extract_equipment_rows
from app.models import Equipment, EquipmentVersion, Project
from app.schemas.common import Page
from app.schemas.equipment import EquipmentCreate, EquipmentOut, EquipmentUpdate
from app.services import audit_service
from app.services.quantity import compute_installed_weight, pick_effective_total
from app.services.version_service import apply_update, record_initial_version

router = APIRouter()


def _equipment_filter_stmt(
    project_id: int,
    *,
    workspace: str | None = None,
    q: str | None = None,
    module: str | None = None,
    equipment_type: str | None = None,
    min_version: int = 1,
    updated_since_hours: int | None = None,
) -> Select:
    """Shared WHERE-clause builder for both the list endpoint (paginated
    display) and the Excel export (exports every matching row regardless
    of page) — keeps the two in sync so "export what you're filtering on"
    is actually correct rather than relying on the client to enumerate ids.
    """
    stmt = select(Equipment).where(Equipment.project_id == project_id)
    if workspace:
        stmt = stmt.where(Equipment.workspace == workspace)
    if q:
        like = f"%{q}%"
        stmt = stmt.where(
            (Equipment.client_tag.ilike(like))
            | (Equipment.old_tag.ilike(like))
            | (Equipment.description.ilike(like))
        )
    if module:
        stmt = stmt.where(Equipment.module == module)
    if equipment_type:
        stmt = stmt.where(Equipment.equipment_type == equipment_type)
    if min_version > 1:
        stmt = stmt.where(Equipment.current_version >= min_version)
    if updated_since_hours is not None:
        cutoff = datetime.now(timezone.utc) - timedelta(hours=updated_since_hours)
        stmt = stmt.where(Equipment.updated_at >= cutoff)
    return stmt


_SORT_COLUMNS = {
    "client_tag": Equipment.client_tag,
    "current_version": Equipment.current_version,
    "updated_at": Equipment.updated_at,
}


@router.get("/projects/{project_id}/equipment", response_model=Page[EquipmentOut])
async def list_equipment(
    db: DbSession,
    project: Project = Depends(project_access("viewer")),
    q: str | None = Query(None, description="Search by client tag, old tag, or description."),
    workspace: str | None = Query(None, description="Filter by workspace: 'topside' or 'marine'. Omit to return all."),
    module: str | None = None,
    equipment_type: str | None = None,
    min_version: int = Query(1, ge=1, description="Only rows with current_version >= this."),
    updated_since_hours: int | None = Query(
        None, ge=1, description="Only rows updated within the last N hours."
    ),
    sort_by: str = Query("client_tag", regex="^(client_tag|current_version|updated_at)$"),
    sort_dir: str = Query("asc", regex="^(asc|desc)$"),
    limit: int = Query(50, ge=1, le=5000),
    offset: int = Query(0, ge=0),
):
    stmt = _equipment_filter_stmt(
        project.id,
        workspace=workspace, q=q, module=module, equipment_type=equipment_type,
        min_version=min_version, updated_since_hours=updated_since_hours,
    )

    total = (
        await db.execute(select(func.count()).select_from(stmt.subquery()))
    ).scalar_one()

    col = _SORT_COLUMNS[sort_by]
    order = col.desc() if sort_dir == "desc" else col.asc()
    stmt = stmt.order_by(order).limit(limit).offset(offset)
    rows = (await db.execute(stmt)).scalars().all()
    return Page(items=rows, total=total, limit=limit, offset=offset)


@router.post("/projects/{project_id}/equipment", response_model=EquipmentOut, status_code=201)
async def create_equipment(
    payload: EquipmentCreate,
    db: DbSession,
    user: CurrentUser,
    project: Project = Depends(project_access("editor")),
):
    data = payload.model_dump()
    extras = data.pop("data", {})
    eq = Equipment(project_id=project.id, data=extras, created_by_id=user.id, **data)
    db.add(eq)
    try:
        await db.flush()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(status_code=400, detail="client_tag already exists for this project")
    await record_initial_version(db, eq, source="manual", user_id=user.id)
    await audit_service.log(
        db, action="equipment.create",
        user_id=user.id, project_id=project.id,
        entity_type="equipment", entity_id=eq.id,
    )
    await db.commit()
    await db.refresh(eq)
    return eq


@router.get("/projects/{project_id}/equipment/{equipment_id}", response_model=EquipmentOut)
async def get_equipment(
    equipment_id: int,
    db: DbSession,
    project: Project = Depends(project_access("viewer")),
):
    eq = (
        await db.execute(
            select(Equipment).where(
                Equipment.id == equipment_id, Equipment.project_id == project.id
            )
        )
    ).scalar_one_or_none()
    if not eq:
        raise HTTPException(status_code=404, detail="Equipment not found")
    return eq


@router.patch("/projects/{project_id}/equipment/{equipment_id}", response_model=EquipmentOut)
async def update_equipment(
    equipment_id: int,
    payload: EquipmentUpdate,
    db: DbSession,
    user: CurrentUser,
    project: Project = Depends(project_access("editor")),
):
    eq = (
        await db.execute(
            select(Equipment).where(
                Equipment.id == equipment_id, Equipment.project_id == project.id
            )
        )
    ).scalar_one_or_none()
    if not eq:
        raise HTTPException(status_code=404, detail="Equipment not found")

    data = payload.model_dump(exclude_unset=True)
    note = data.pop("note", None)
    extras: dict[str, Any] | None = data.pop("data", None)
    await apply_update(
        db, eq, data,
        source="manual",
        source_file_id=None,
        user_id=user.id,
        note=note,
        extra_data=extras,
    )
    await audit_service.log(
        db, action="equipment.update",
        user_id=user.id, project_id=project.id,
        entity_type="equipment", entity_id=eq.id,
        metadata={"changes": list(data.keys())},
    )
    await db.commit()
    await db.refresh(eq)
    return eq


@router.delete("/projects/{project_id}/equipment/{equipment_id}", status_code=204)
async def delete_equipment(
    equipment_id: int,
    db: DbSession,
    user: CurrentUser,
    project: Project = Depends(project_access("editor")),
):
    eq = (
        await db.execute(
            select(Equipment).where(
                Equipment.id == equipment_id, Equipment.project_id == project.id
            )
        )
    ).scalar_one_or_none()
    if not eq:
        raise HTTPException(status_code=404, detail="Equipment not found")
    await audit_service.log(
        db, action="equipment.delete",
        user_id=user.id, project_id=project.id,
        entity_type="equipment", entity_id=eq.id,
    )
    await db.delete(eq)
    await db.commit()


class BulkDeleteRequest(BaseModel):
    ids: list[int] = Field(..., min_length=1, max_length=5000)


@router.post("/projects/{project_id}/equipment/bulk-delete")
async def bulk_delete_equipment(
    payload: BulkDeleteRequest,
    db: DbSession,
    user: CurrentUser,
    project: Project = Depends(project_access("editor")),
):
    """Delete many equipment rows in one transaction.

    Only rows that actually belong to ``project_id`` are deleted — IDs
    from other projects (or non-existent IDs) are silently ignored, so a
    stale UI never wipes the wrong rows. Returns counts the caller can
    show in a toast: ``deleted`` (actual rows removed) + ``not_found``
    (ids the user asked about that weren't in this project).
    """
    if not payload.ids:
        return {"deleted": 0, "not_found": 0}

    rows = (
        await db.execute(
            select(Equipment).where(
                Equipment.id.in_(payload.ids),
                Equipment.project_id == project.id,
            )
        )
    ).scalars().all()
    found_ids = {r.id for r in rows}
    not_found = len([i for i in payload.ids if i not in found_ids])

    for r in rows:
        await db.delete(r)

    await audit_service.log(
        db,
        action="equipment.bulk_delete",
        user_id=user.id,
        project_id=project.id,
        metadata={
            "deleted": len(found_ids),
            "not_found": not_found,
            "ids": list(found_ids),
        },
    )
    await db.commit()
    return {"deleted": len(found_ids), "not_found": not_found}


# ---------------------------------------------------------------------------
# Excel export — matches the reference EPC template
# (``20171-SPOG-80000-ME-LS-0001_Z1_Topside Eqipment List.xlsx``)
# ---------------------------------------------------------------------------
#
# Each tuple: (attr_on_Equipment, header_text_with_explicit_newlines, width).
# Header text uses real "\n" so cells with wrap_text=True render multi-line
# (e.g. "EQUIPMENT\nTYPE", "OPERATING PRESS\n(barg)") — same as the reference.
# Order, casing, spacing, and trailing-spaces are preserved verbatim from the
# reference workbook so the downstream EPC workflow that consumes this file
# doesn't see any surprises.
EXPORT_COLUMNS: list[tuple[str, str, int]] = [
    ("rev_no",                    "REV No.",                              8),
    ("old_tag",                   "OLD 'EQUIPMENT /\nTAG No.",            15),
    ("client_tag",                "CLIENT EQUIPMENT \nTAG",               18),
    ("description",               "DESCRIPTION",                          36),
    ("vendor",                    "VENDOR",                               22),
    ("equipment_type",            "EQUIPMENT\nTYPE",                      18),
    ("module",                    "MODULE",                               12),
    ("design_code",               "EQUIPMENT\nDESIGN\nCODE/CLASS",        16),
    ("orientation",               "ORIENTATION",                          12),
    ("material",                  "MATERIAL OF CONSTRUCTION ",            26),
    ("configuration",             "CONFIGURATION",                        14),
    ("location",                  "LOCATION",                             18),
    ("operating_press",           "OPERATING PRESS\n(barg)",              14),
    ("operating_temp",            "OPERATING TEMP\n(oC)",                 14),
    ("design_press",              "DESIGN PRESS\n(barg)",                 12),
    ("design_temp",               "DESIGN TEMP\n(oC)",                    12),
    ("design_flow",               "DESIGN FLOW m3/hr",                    14),
    ("pump_capacity",             "PUMP / COMPRESSOR / TANK CAPACITY",    20),
    ("heat_exchanger_duty_kw",    " HEAT EXCHANGER DUTY (kW)",            18),
    ("liquid_fill",               "LIQUID FILL",                          12),
    ("absorbed_power_kw",         "ABSORBED\nPOWER PER UNIT\n(kW)",       14),
    ("rated_power_kw",            "RATED\nPOWER PER UNIT\n(kW)",          14),
    ("length_m",                  "L or\nT/T\n(m)",                       9),
    ("width_id_m",                "W or I.D\n(m)",                        9),
    ("height_tt_m",               "H or T/T\n(m)",                        9),
    ("dry_weight_mt",             "DRY WT       in MT",                   12),
    ("operating_weight_mt",       "OPE WT in MT",                         12),
    ("hydrotest_weight_mt",       "HYDROTEST WT\nin MT",                  14),
    ("pid",                       "P&ID",                                 24),
    ("remarks",                   "REMARKS",                              22),
    ("total_dry_weight_mt",       "TOTAL DRY WT\nin mT",                  12),
    ("total_operating_weight_mt", "TOTAL OPE WT\nin mT",                  12),
    ("lifecycle_status",          "LIFECYCLE",                            14),
    # Extra fields from vendor drawings — appended after the reference-
    # matching layout so the first 31 columns stay identical to the
    # contractor's template.
    ("length_overall_m",          "OVERALL LENGTH\n(m)",                  12),
    ("mdmt_c",                    "MDMT\n(oC)",                           10),
    ("hydrostatic_test_press_barg", "HYDRO TEST PRESS\n(barg)",           14),
    ("insulation",                "INSULATION",                           22),
]

# Header layout — mirrors the reference EPC workbook
# (`20171-SPOG-80000-ME-LS-0001_Z1_Topside Eqipment List.xlsx`).
#
#   ┌──────────────┬─────────── TOPSIDES EQUIPMENT LIST ────────┬──────────┐
#   │              │                                            │ COMPANY  │
#   │  LOGO AREA   │                                            │ PROJECT  │
#   │  (reserved)  │                                            │ PROJECT  │
#   │              │                                            │ Doc No.  │
#   └──────────────┴─────────────────────────────────────────────┴──────────┘
#   │   REV │ OLD TAG │ CLIENT TAG │ … (column headers with filter arrows) │
#   ├──────────────────────────────────────────────────────────────────────┤
#   │ MODULE MD: FLARE KNOCK OUT DRUMS & PUMPS / …  (section banner)       │
#   │ <equipment rows>                                                      │
#   ├──────────────────────────────────────────────────────────────────────┤
#   │ MODULE M10: …                                                         │
#   │ <equipment rows>                                                      │
#   └──────────────────────────────────────────────────────────────────────┘
_LOGO_COL_START = 1   # column A — reserved logo area (left side)
_LOGO_COL_END = 6     # column F
_TITLE_COL_START = 7  # column G — "TOPSIDES EQUIPMENT LIST" centered
_TITLE_COL_END = 24   # column X
_TITLE_LABEL_COL = 26 # column Z   — "COMPANY", "PROJECT No.", …
_TITLE_LABEL_END = 28 # column AB  — merged label cell
_TITLE_VALUE_COL = 29 # column AC  — "ONGC", "20171", …
_TITLE_VALUE_END = 31 # column AE  — merged value cell to right edge
_TITLE_BLOCK_ROWS = 4 # rows 1-4

_HEADER_FILL = PatternFill("solid", fgColor="FFF2CC")  # pale yellow — column headers
_BANNER_FILL = PatternFill("solid", fgColor="C6EFCE")  # mint green   — module banners
_TITLE_LABEL_FILL = PatternFill("solid", fgColor="F2F2F2")  # light grey — title-block labels
# Version-rank highlighting: cells whose value came from the CURRENT
# equipment version get GREEN; cells last touched by the PREVIOUS
# version get ORANGE. Older changes and initial-creation cells stay
# white so the sheet only calls attention to the two most recent moves.
_LATEST_FILL = PatternFill("solid", fgColor="C6EFCE")  # Excel "Good"    — green
_LATEST_FONT_COLOR = "006100"
_SECOND_FILL = PatternFill("solid", fgColor="FFCC99")  # soft orange
_SECOND_FONT_COLOR = "974706"
_THIN  = Side(style="thin",  color="999999")
_MED   = Side(style="medium", color="000000")
_BORDER     = Border(top=_THIN, left=_THIN, right=_THIN, bottom=_THIN)
_BOX_BORDER = Border(top=_MED,  left=_MED,  right=_MED,  bottom=_MED)


def _value_for_export(v: Any) -> Any:
    """Render an equipment value for Excel.

    Numbers stored as strings stay as strings (the reference workbook
    routinely uses ``"38.0"`` / ``"FV / 7"`` / ``"-29 / 120 "`` in numeric-
    looking columns). None becomes the literal ``"-"`` the reference uses
    for empty cells, so a downstream parser doesn't see blank cells where
    the contractor expects a dash.
    """
    if v is None or v == "":
        return "-"
    return v


def _find_header_logo() -> Path | None:
    """Look for a header image to embed in the Excel export's logo area.

    Checks ``app/static/logos/`` for the first file matching one of the
    accepted names. We try the project's branded filename first
    (``SP-Oil-Gas.png``, same as the frontend uses at
    ``public/images/SP-Oil-Gas.png``) so the two stay visually
    synchronised; then we fall back to a generic ``header.*`` name for
    setups that prefer a single drop-in slot. Returning ``None`` is the
    normal "no logo configured" case — the export shows a "[ Logos ]"
    text placeholder instead.
    """
    base = Path(__file__).resolve().parent.parent.parent / "static" / "logos"
    candidates = (
        "SP-Oil-Gas.png", "SP-Oil-Gas.jpg", "SP-Oil-Gas.jpeg",
        "header.png", "header.jpg", "header.jpeg",
    )
    for name in candidates:
        p = base / name
        if p.exists() and p.is_file():
            return p
    return None


_SOURCE_LABEL = {
    "manual": "Manual edit",
    "excel":  "Excel import",
    "pfd":    "PFD sync",
    "pid":    "P&ID sync",
    "vendor": "Vendor sync",
    "seed":   "Initial seed",
}


def _make_change_comment(change: dict[str, Any]) -> Comment:
    """Build the hover-comment for a highlighted cell.

    ``change`` is the dict we stashed in ``highlight_map`` — the latest
    ``EquipmentVersion`` row (within the highlight window) that touched
    the field this cell renders. Example text::

        Changed in v3
        Source: Vendor sync
        When:   2026-06-30 14:20 UTC

    Author is set to "MEL" so the Excel comment sidebar labels these
    consistently, distinct from any comments a human might add after
    download.
    """
    when = change.get("created_at")
    when_str = when.strftime("%Y-%m-%d %H:%M UTC") if isinstance(when, datetime) else "—"
    source_raw = str(change.get("source") or "").lower()
    source_label = _SOURCE_LABEL.get(source_raw, source_raw or "—")
    text = (
        f"Changed in v{change.get('version_no', '?')}\n"
        f"Source: {source_label}\n"
        f"When:   {when_str}"
    )
    c = Comment(text, "MEL")
    # Give the comment enough real estate that the three lines don't wrap
    # awkwardly the first time a user hovers. Excel treats these as pt.
    c.width = 240
    c.height = 80
    return c


def _document_no(project: Project, workspace: str) -> str:
    """Best-effort document number for the title block.

    We don't know the contractor's exact numbering scheme, so we fall back
    to ``<code> — <WORKSPACE> EQUIPMENT LIST``. Users can overwrite the
    cell after download if they need the precise EPC document number.
    """
    code = (project.code or "").strip()
    label = f"{workspace.upper()} EQUIPMENT LIST"
    return f"{code} — {label}" if code else label


class ExportFilter(BaseModel):
    """Optional payload for the Excel export.

    ``ids`` (legacy): restrict to an explicit id set the frontend already
    enumerated client-side.

    The remaining fields mirror ``list_equipment``'s filters — used so the
    UI can export "everything matching these filters" even though the
    on-screen table itself is paginated (only one page of rows is ever
    loaded client-side, so there's no full id list to send anymore).
    When ``ids`` is given it wins; otherwise these filters apply.
    """
    ids: list[int] | None = None
    q: str | None = None
    min_version: int = 1
    updated_since_hours: int | None = None


@router.post("/projects/{project_id}/export/excel")
async def export_excel(
    db: DbSession,
    project: Project = Depends(project_access("viewer")),
    workspace: str | None = Query(
        None,
        regex="^(topside|marine)$",
        description="Restrict the export to one workspace.",
    ),
    highlight: bool = Query(
        True,
        description=(
            "If true (default), cells set by the current equipment version "
            "are filled green and cells last set by the previous version "
            "are filled orange. Older changes stay white. Set false for a "
            "pristine EPC-format export with no coloring."
        ),
    ),
    body: ExportFilter | None = Body(default=None),
):
    """Download the project's equipment list as an .xlsx in the EPC
    reference layout — title block top-right, multi-line bold headers,
    module section banners, the 31-column structure identical to the
    source template the contractor delivered.

    POST (not GET) because the optional ``ids`` filter can carry hundreds
    of integers — well past safe URL-length limits for a GET request.
    The body is optional; calling with no body exports everything in
    the (project, workspace) scope.
    """
    if body and body.ids:
        # Restrict to the explicit ID set the frontend sent. We keep the
        # project + workspace filter on top so a user can't smuggle in IDs
        # they don't have access to via project_access().
        stmt = select(Equipment).where(
            Equipment.project_id == project.id, Equipment.id.in_(body.ids)
        )
        if workspace:
            stmt = stmt.where(Equipment.workspace == workspace)
    else:
        stmt = _equipment_filter_stmt(
            project.id,
            workspace=workspace,
            q=body.q if body else None,
            min_version=body.min_version if body else 1,
            updated_since_hours=body.updated_since_hours if body else None,
        )
    # Sort by module first so we can drop a banner row at the start of
    # each group, then by client_tag inside each module — same order
    # the reference workbook uses.
    stmt = stmt.order_by(Equipment.module.asc().nullslast(), Equipment.client_tag.asc())
    rows = (await db.execute(stmt)).scalars().all()

    # Version-rank highlighting.
    #
    # Per equipment, we identify the two most-recent NON-INITIAL versions
    # (version_no >= 2 — the seed / first-create snapshot at v1 has
    # `changed_fields = ALL_TRACKED_FIELDS`, so counting it would turn
    # every field on a fresh row green and defeat the point).
    #
    # For each equipment we build:
    #   latest_change[eq_id] -> {version_no, source, created_at, fields}
    #   second_change[eq_id] -> {…} or absent
    #
    # In the cell loop we colour a cell:
    #   • GREEN  if its attr is in latest_change[eq.id]["fields"]
    #   • ORANGE elif its attr is in second_change[eq.id]["fields"]
    #   • no fill otherwise
    #
    # A field that changed in v2, then again in v4 — with v5 being
    # current but not touching it — will show up as ORANGE (last actual
    # move to that value was v4, which is the previous non-initial
    # version) rather than green. That matches the user's intuition:
    # "orange = one behind the newest change".
    latest_change: dict[int, dict[str, Any]] = {}
    second_change: dict[int, dict[str, Any]] = {}
    if highlight and rows:
        v_stmt = (
            select(
                EquipmentVersion.equipment_id,
                EquipmentVersion.version_no,
                EquipmentVersion.changed_fields,
                EquipmentVersion.source,
                EquipmentVersion.created_at,
            )
            .where(
                EquipmentVersion.equipment_id.in_([r.id for r in rows]),
                EquipmentVersion.version_no >= 2,
            )
            .order_by(
                EquipmentVersion.equipment_id.asc(),
                EquipmentVersion.version_no.desc(),  # newest first per equipment
            )
        )
        for eq_id, ver_no, changed, source, created_at in (await db.execute(v_stmt)).all():
            if not changed:
                continue
            entry = {
                "version_no": ver_no,
                "source": source,
                "created_at": created_at,
                "fields": set(changed),
            }
            if eq_id not in latest_change:
                latest_change[eq_id] = entry
            elif eq_id not in second_change:
                second_change[eq_id] = entry
            # else: already have the top two; keep scanning is cheap and
            # short-circuiting per-eq isn't worth the added complexity.

    workspace_label = (workspace or project.project_type or "topside").lower()
    is_marine = workspace_label == "marine"
    title = "MARINE EQUIPMENT LIST" if is_marine else "TOPSIDES EQUIPMENT LIST"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EQUIPMENT LIST"
    n_cols = len(EXPORT_COLUMNS)

    # ── Title block (rows 1-4) ──────────────────────────────────────────
    # Three regions on the same 4-row band:
    #   1. Left:   reserved logo area (rows 1-4, columns A-F). Empty, but
    #              bordered + sized so a project admin can paste in the
    #              client / contractor / FPSO operator logos after download.
    #   2. Centre: "TOPSIDES EQUIPMENT LIST" title, merged across rows 1-4
    #              and columns G-X. Big, bold, centered.
    #   3. Right:  Project metadata (COMPANY, PROJECT No., PROJECT, Doc No),
    #              one row each, with label cells in Z:AB and value cells
    #              in AC:AE. Bordered.

    # 1) Logo area — merge + border. We attempt to embed a header image
    # found under app/static/logos/. If none exists we leave the area
    # blank with a "[ Logos ]" placeholder so the layout still looks
    # right; just drop a PNG/JPG into that folder and every future
    # export picks it up automatically.
    ws.merge_cells(
        start_row=1, start_column=_LOGO_COL_START,
        end_row=_TITLE_BLOCK_ROWS, end_column=_LOGO_COL_END,
    )
    for r in range(1, _TITLE_BLOCK_ROWS + 1):
        for c in range(_LOGO_COL_START, _LOGO_COL_END + 1):
            ws.cell(row=r, column=c).border = _BOX_BORDER
    logo_path = _find_header_logo()
    if logo_path is not None:
        try:
            img = XLImage(str(logo_path))
            # Scale to a fixed target HEIGHT that fits the 4-row title
            # block, preserving the logo's natural aspect ratio so it
            # doesn't get visually stretched. openpyxl sets `img.width`
            # and `img.height` from the file by default, so we read
            # those and rescale.
            natural_w = img.width or 1
            natural_h = img.height or 1
            target_h = 90  # px — fits inside the ~110px title-block area
            target_w = int(target_h * natural_w / natural_h)
            img.width = target_w
            img.height = target_h
            # Anchor near the top-left of A1 with a small offset so the
            # logo doesn't touch the cell borders.
            img.anchor = "A1"
            ws.add_image(img, "A1")
        except Exception:
            # openpyxl raises on unreadable / unsupported images. Fall
            # back to the placeholder rather than failing the whole
            # export.
            logo_path = None
    if logo_path is None:
        placeholder = ws.cell(row=1, column=_LOGO_COL_START)
        placeholder.value = "[ Logos ]"
        placeholder.font = Font(italic=True, size=9, color="999999")
        placeholder.alignment = Alignment(horizontal="center", vertical="center")

    # 2) Centered title — merged
    ws.merge_cells(
        start_row=1, start_column=_TITLE_COL_START,
        end_row=_TITLE_BLOCK_ROWS, end_column=_TITLE_COL_END,
    )
    title_cell = ws.cell(row=1, column=_TITLE_COL_START)
    title_cell.value = title
    title_cell.font = Font(bold=True, size=18, color="1F4E78")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in range(1, _TITLE_BLOCK_ROWS + 1):
        for c in range(_TITLE_COL_START, _TITLE_COL_END + 1):
            ws.cell(row=r, column=c).border = _BOX_BORDER

    # 3) Project metadata block (top-right)
    title_block = [
        (1, "COMPANY",      project.client or ""),
        (2, "PROJECT No.",  project.code or ""),
        (3, "PROJECT",      project.name or ""),
        (4, "Document No",  _document_no(project, workspace_label)),
    ]
    for row, label, value in title_block:
        # Label (merged Z:AB)
        ws.merge_cells(
            start_row=row, start_column=_TITLE_LABEL_COL,
            end_row=row, end_column=_TITLE_LABEL_END,
        )
        lc = ws.cell(row=row, column=_TITLE_LABEL_COL, value=label)
        lc.font = Font(bold=True, size=10)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lc.fill = _TITLE_LABEL_FILL
        for c in range(_TITLE_LABEL_COL, _TITLE_LABEL_END + 1):
            ws.cell(row=row, column=c).border = _BORDER
        # Value (merged AC:AE)
        ws.merge_cells(
            start_row=row, start_column=_TITLE_VALUE_COL,
            end_row=row, end_column=_TITLE_VALUE_END,
        )
        vc = ws.cell(row=row, column=_TITLE_VALUE_COL, value=value)
        vc.font = Font(size=10)
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for c in range(_TITLE_VALUE_COL, _TITLE_VALUE_END + 1):
            ws.cell(row=row, column=c).border = _BORDER

    # Make the title band tall enough to breathe (matches the reference's
    # ~30pt-per-row look — 4 rows × 28 = 112pt total).
    for r in range(1, _TITLE_BLOCK_ROWS + 1):
        ws.row_dimensions[r].height = 28

    # ── Column headers (row 5) — multi-line, bold, filled, bordered ────
    HEADER_ROW = 5
    header_font = Font(bold=True, size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for idx, (_, header_text, width) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=idx, value=header_text)
        cell.font = header_font
        cell.alignment = header_align
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[HEADER_ROW].height = 52

    # AutoFilter on the column-header row — this is what gives every
    # header the little dropdown arrow visible in the reference workbook
    # (Data → Filter in Excel). The range covers the full data area;
    # openpyxl extends it automatically when we add rows below.
    last_col_letter = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A{HEADER_ROW}:{last_col_letter}{HEADER_ROW}"

    # ── Data rows with module section banners ───────────────────────────
    # `n_cols` was already computed above (it determines auto_filter.ref).
    body_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    banner_align = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    banner_font = Font(bold=True, size=11, color="006100")  # dark green on mint
    body_font = Font(size=10)

    current_row = HEADER_ROW + 1
    last_module: str | None = None
    for eq in rows:
        module = (eq.module or "—").strip()
        if module != last_module:
            # New module group — merged banner row spanning all data
            # columns, like the reference does at the start of each
            # section (e.g. "MD: FLARE KNOCK OUT DRUMS & PUMPS …").
            banner_cell = ws.cell(row=current_row, column=1, value=f"MODULE {module}")
            banner_cell.font = banner_font
            banner_cell.alignment = banner_align
            banner_cell.fill = _BANNER_FILL
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=n_cols,
            )
            for c in range(1, n_cols + 1):
                ws.cell(row=current_row, column=c).border = _BORDER
            current_row += 1
            last_module = module

        latest = latest_change.get(eq.id)
        second = second_change.get(eq.id)
        latest_fields: set[str] = latest["fields"] if latest else set()
        second_fields: set[str] = second["fields"] if second else set()
        # Fill missing TOTAL WT columns from DRY/OPE × count(CONFIGURATION).
        # The reference EPC template treats DRY WT as per-unit weight and
        # TOTAL DRY WT as the installed weight (per-unit × units-in-config,
        # e.g. "2 x 100%" → ×2). When the DB has DRY but no TOTAL, we
        # substitute the computed value so the workbook always has a
        # sensible TOTAL column for downstream consumers.
        computed_total_dry = compute_installed_weight(
            eq.dry_weight_mt, eq.configuration,
        )
        computed_total_ope = compute_installed_weight(
            eq.operating_weight_mt, eq.configuration,
        )
        effective_total_dry = pick_effective_total(
            eq.total_dry_weight_mt, computed_total_dry,
        )
        effective_total_ope = pick_effective_total(
            eq.total_operating_weight_mt, computed_total_ope,
        )
        for idx, (attr, _, _) in enumerate(EXPORT_COLUMNS, start=1):
            if attr == "total_dry_weight_mt":
                raw = effective_total_dry
            elif attr == "total_operating_weight_mt":
                raw = effective_total_ope
            else:
                raw = getattr(eq, attr) if attr else None
            val = _value_for_export(raw)
            cell = ws.cell(row=current_row, column=idx, value=val)
            cell.alignment = body_align
            cell.border = _BORDER
            # GREEN if the current version touched this field; ORANGE
            # if it wasn't in the current version but was in the one
            # before it; no fill otherwise. Hover-comment names the
            # exact version that owns each colour.
            if attr and attr in latest_fields:
                cell.fill = _LATEST_FILL
                cell.font = Font(size=10, bold=True, color=_LATEST_FONT_COLOR)
                cell.comment = _make_change_comment(latest)
            elif attr and attr in second_fields:
                cell.fill = _SECOND_FILL
                cell.font = Font(size=10, bold=True, color=_SECOND_FONT_COLOR)
                cell.comment = _make_change_comment(second)
            else:
                cell.font = body_font
        current_row += 1

    # Freeze the header band so scrolling keeps it visible while reviewing
    # the equipment rows underneath.
    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{project.code or 'project'}_{workspace_label}_equipment_list.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# --- Bulk import from Excel ---

# Fields that can be set from a parsed Excel row (mirrors EquipmentCreate).
_IMPORTABLE_FIELDS = [
    "rev_no", "old_tag", "client_tag", "description", "vendor", "equipment_type",
    "module", "design_code", "orientation", "material", "configuration", "location",
    "operating_press", "operating_temp", "design_press", "design_temp", "design_flow",
    "pump_capacity", "heat_exchanger_duty_kw", "liquid_fill",
    "absorbed_power_kw", "rated_power_kw",
    "length_m", "width_id_m", "height_tt_m",
    "dry_weight_mt", "operating_weight_mt", "hydrotest_weight_mt",
    "pid", "remarks", "total_dry_weight_mt", "total_operating_weight_mt",
    "lifecycle_status",
    "length_overall_m", "mdmt_c", "hydrostatic_test_press_barg", "insulation",
]


@router.post("/projects/{project_id}/equipment/import")
async def import_equipment_excel(
    db: DbSession,
    user: CurrentUser,
    file: UploadFile = File(..., description="Equipment List .xlsx file"),
    sheet_name: str | None = Query(None, description="Override sheet to read; default = 'EQUIPMENT LIST' or the largest sheet."),
    commit: bool = Query(False, description="If false, return a parse preview without writing to the database."),
    mode: str = Query(
        "skip_existing",
        regex="^(skip_existing|update_existing)$",
        description="Conflict policy when commit=true: skip_existing leaves matched rows alone; update_existing PATCHes them.",
    ),
    workspace: str = Query(
        "topside",
        regex="^(topside|marine)$",
        description="Which workspace these rows belong to: 'topside' or 'marine'.",
    ),
    project: Project = Depends(project_access("editor")),
):
    """Bulk-import equipment from a Topside-Equipment-List style Excel file.

    Same parser as scripts/seed_topside_poc.py — locates the EQUIPMENT row
    header by content patterns, builds a column map, then walks the data rows.

    With `commit=false` (default) the response is a preview only; the client
    can show the user what would be imported, then re-POST with `commit=true`.
    """
    # Persist the upload to a tempfile because openpyxl needs a real path.
    suffix = Path(file.filename or "").suffix.lower() or ".xlsx"
    if suffix not in (".xlsx", ".xlsm"):
        raise HTTPException(status_code=400, detail="Only .xlsx / .xlsm are supported")

    body = await file.read()
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(body)
        tmp_path = tmp.name

    try:
        try:
            parsed = extract_equipment_rows(tmp_path, sheet_name=sheet_name)
        except Exception as e:  # noqa: BLE001
            raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {e}")
    finally:
        try:
            Path(tmp_path).unlink(missing_ok=True)
        except Exception:
            pass

    if not parsed:
        raise HTTPException(
            status_code=400,
            detail="No equipment rows recognized in the file. Make sure the sheet has a 'CLIENT EQUIPMENT TAG' header.",
        )

    # Look up existing client_tags WITHIN THIS WORKSPACE only. Topsides and
    # Marine can independently have a row called e.g. "P-F22001A/B" without
    # collision; we treat them as separate namespaces.
    existing_rows = (
        await db.execute(
            select(Equipment).where(
                Equipment.project_id == project.id,
                Equipment.workspace == workspace,
            )
        )
    ).scalars().all()
    existing_by_tag = {r.client_tag: r for r in existing_rows}

    # Track tags we've already seen IN THIS FILE so a duplicate inside the
    # Excel itself (same tag repeated on row 47 and row 122, say) doesn't
    # explode the import with a uniqueness violation. We keep the FIRST
    # occurrence and mark every subsequent one as ``duplicate_in_file`` so
    # the user can see exactly which rows collided and decide whether to
    # clean up the source spreadsheet.
    seen_in_file: dict[str, int] = {}  # tag -> first row_number it appeared on

    preview: list[dict[str, Any]] = []
    for i, row in enumerate(parsed, start=1):
        tag = (row.get("client_tag") or "").strip()
        if not tag:
            preview.append({
                "row_number": i, "client_tag": None,
                "status": "invalid", "reason": "missing client_tag",
            })
            continue

        # Intra-file dedup. The DB-level uniqueness is
        # (project, workspace, client_tag), so two rows in the same Excel
        # with the same tag can't both land — we keep the first.
        first_row = seen_in_file.get(tag)
        if first_row is not None:
            preview.append({
                "row_number": i,
                "client_tag": tag,
                "status": "duplicate_in_file",
                "reason": f"Same tag also appeared earlier on row {first_row}",
            })
            continue
        seen_in_file[tag] = i

        is_existing = tag in existing_by_tag
        clean: dict[str, Any] = {}
        for k in _IMPORTABLE_FIELDS:
            v = row.get(k)
            if isinstance(v, str):
                v = v.strip()
                if not v or v == "-":
                    v = None
            clean[k] = v
        preview.append({
            "row_number": i,
            "client_tag": tag,
            "status": "existing" if is_existing else "new",
            "fields": clean,
            "raw_extra": row.get("__raw") or {},
        })

    summary = {
        "total_rows": len(parsed),
        "new": sum(1 for p in preview if p["status"] == "new"),
        "existing": sum(1 for p in preview if p["status"] == "existing"),
        "invalid": sum(1 for p in preview if p["status"] == "invalid"),
        "duplicate_in_file": sum(1 for p in preview if p["status"] == "duplicate_in_file"),
        "commit": commit,
        "mode": mode,
    }

    if not commit:
        # Return preview only — first 200 rows to keep response sane
        return {
            **summary,
            "preview": preview[:200],
            "preview_truncated": len(preview) > 200,
        }

    # Commit path. SAVEPOINT-per-row was costing us 4 DB round-trips per row
    # (~13 minutes on 668 rows over the Oregon → India Render link). We
    # instead:
    #   1. Build all NEW Equipment objects in memory, bulk add_all + one
    #      flush so the DB does a single multi-row INSERT.
    #   2. Build all v1 EquipmentVersion objects (now that we have IDs from
    #      step 1) and bulk add_all + one flush.
    #   3. Existing-row updates still use apply_update so the per-field
    #      precedence + range-preservation rules apply, but without the
    #      SAVEPOINT — one bad existing row would only affect itself, the
    #      rest still commit because of the outer transaction's atomicity
    #      around the new-row INSERTs which already succeeded.
    created = 0
    updated = 0
    skipped = 0
    errors: list[dict[str, Any]] = []

    # ---- Pass 1: build the new-row equipment objects --------------------
    from app.models import EquipmentVersion  # local import to keep module top tidy
    from app.services.version_service import snapshot, TRACKED_FIELDS

    new_eq_objects: list[Equipment] = []
    new_eq_row_numbers: list[int] = []
    for p in preview:
        if p["status"] != "new":
            continue
        tag = p["client_tag"]
        fields = p.get("fields") or {}
        raw_extra = p.get("raw_extra") or {}
        try:
            eq = Equipment(
                project_id=project.id,
                workspace=workspace,
                data={"raw": raw_extra},
                created_by_id=user.id,
                current_version=1,
                last_source="excel",
                last_updated_by_id=user.id,
                **fields,
            )
            new_eq_objects.append(eq)
            new_eq_row_numbers.append(p["row_number"])
        except Exception as e:  # noqa: BLE001
            errors.append({"row_number": p["row_number"], "tag": tag, "error": str(e)})

    if new_eq_objects:
        try:
            db.add_all(new_eq_objects)
            await db.flush()   # one INSERT ... VALUES (...), (...), ... ; one round-trip
            # Build all v1 versions in memory now that equipment_ids exist.
            versions = [
                EquipmentVersion(
                    equipment_id=eq.id,
                    version_no=1,
                    snapshot=snapshot(eq),
                    changed_fields=list(TRACKED_FIELDS),
                    source="excel",
                    source_file_id=None,
                    created_by_id=user.id,
                )
                for eq in new_eq_objects
            ]
            db.add_all(versions)
            await db.flush()
            created = len(new_eq_objects)
        except IntegrityError as e:
            await db.rollback()
            raise HTTPException(
                status_code=409,
                detail=(
                    "Duplicate client_tag(s) within this workspace caused a "
                    f"constraint violation. The uniqueness key is "
                    f"(project, workspace, client_tag): {e.orig}"
                ),
            )

    # ---- Pass 2: update existing rows (only if mode allows) --------------
    if mode == "update_existing":
        for p in preview:
            if p["status"] != "existing":
                continue
            tag = p["client_tag"]
            fields = p.get("fields") or {}
            raw_extra = p.get("raw_extra") or {}
            try:
                eq = existing_by_tag[tag]
                changes = {
                    k: v for k, v in fields.items()
                    if k != "client_tag" and v is not None
                }
                v = await apply_update(
                    db, eq, changes,
                    source="excel",
                    source_file_id=None,
                    user_id=user.id,
                    note=f"Imported from {file.filename}",
                    extra_data={"raw": raw_extra} if raw_extra else None,
                )
                if v:
                    updated += 1
                else:
                    skipped += 1
            except Exception as e:  # noqa: BLE001
                errors.append({"row_number": p["row_number"], "tag": tag, "error": str(e)})
    else:
        # skip_existing mode — just count
        skipped += sum(1 for p in preview if p["status"] == "existing")

    # Invalid rows always count as skipped regardless of mode
    skipped += sum(1 for p in preview if p["status"] == "invalid")

    # Intra-file duplicates are surfaced in `errors` so the user sees the
    # exact row numbers that collided, plus added to `skipped` so the
    # totals add up to `total_rows`.
    for p in preview:
        if p["status"] == "duplicate_in_file":
            skipped += 1
            errors.append({
                "row_number": p["row_number"],
                "tag": p["client_tag"],
                "error": p.get("reason") or "duplicate tag in file",
            })

    try:
        await audit_service.log(
            db,
            action="equipment.import_excel",
            user_id=user.id,
            project_id=project.id,
            metadata={
                "filename": file.filename,
                "total_rows": summary["total_rows"],
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors": len(errors),
                "mode": mode,
            },
        )
        await db.commit()
    except Exception as e:  # noqa: BLE001
        await db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Failed to commit import: {e}",
        )

    return {
        **summary,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    }
