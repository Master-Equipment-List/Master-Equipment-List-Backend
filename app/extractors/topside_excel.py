"""Equipment List Excel importer — handles BOTH Topsides and Marine formats.

Tolerant of the layout variations between the project's reference workbooks:
  - **Topsides MEL** (20171-SPOG / 40801-SPE): "CLIENT EQUIPMENT TAG",
    "DESCRIPTION", separate L/W/H columns, etc.
  - **Marine MEL** (20171-SPOG-60000-MA): "TAG NUMBER",
    "EQUIPMENT DESCRIPTION", a single combined "DIMENSION (mm) L x W x H"
    column, "BRAND/MAKER" instead of "VENDOR", "SYSTEM CODE" instead of
    "MODULE", etc.

We auto-detect the header row, build a column map from a generous alias
list, then parse the combined-dimensions cell when present.
"""
from __future__ import annotations

import re
from typing import Any

import openpyxl

# Normalized field key -> list of header substrings (case-insensitive).
# The first matching column wins. Both Topsides and Marine aliases live
# here together so the same importer handles either workbook.
FIELD_HEADER_PATTERNS: dict[str, list[str]] = {
    "rev_no": ["rev no", "rev no.", "rev."],
    "old_tag": ["old equipment", "old tag", "old 'equipment", "old equip"],
    "client_tag": [
        # Topsides
        "client equipment", "tag no.", "tag no", "client tag",
        # Marine
        "tag number", "tag_number", "equipment tag",
    ],
    "description": [
        "description",            # matches "DESCRIPTION" + "EQUIPMENT DESCRIPTION"
        "equipment description",  # explicit Marine header
    ],
    "vendor": [
        "vendor",
        "brand/maker", "brand / maker", "brand /maker", "brand/ maker",
        "brand", "maker", "manufacturer",
    ],
    "equipment_type": ["equipment type", "equipment\ntype"],
    "module": [
        "module",
        # Marine uses "SYSTEM CODE" / "SYSTEM / CODE" for what Topsides calls module
        "system code", "system / code", "system/code", "system code.",
        "system",
    ],
    "design_code": ["design code", "design\ncode/class", "code/class"],
    "orientation": ["orientation"],
    "material": ["material"],
    "configuration": ["configuration", "quantity and configuration"],
    "location": ["location"],
    "operating_press": ["operating pressure", "operating press"],
    "operating_temp": ["operating temperature", "operating temp"],
    "design_press": ["design pressure", "design press"],
    "design_temp": ["design temperature", "design temp"],
    "design_flow": ["design flow"],
    "pump_capacity": [
        "pump / compressor", "pump/compressor", "pump capacity", "tank capac",
        "capacity",  # bare fallback — last, so more specific matches above win
    ],
    "heat_exchanger_duty_kw": ["heat exchanger duty"],
    "liquid_fill": ["liquid fill"],
    "absorbed_power_kw": [
        "absorbed power", "absorbed\npower",
        "electrical absorbed power",  # Marine
        "power per unit",  # Marine: single unqualified power column
    ],
    "rated_power_kw": ["rated power", "rated\npower"],
    # "l (m)" / "w (m)" / "h (m)" are the Marine spelling for separate
    # L/W/H columns (Topsides uses "L OR\nT/T" etc). Order matters here:
    # these must NOT be added to the combined-dimension patterns below.
    "length_m": ["l or", "l or\nt/t", "l (m)"],
    "width_id_m": ["w or", "w or\ni.d", "w (m)"],
    "height_tt_m": ["h or", "h or\nt/t", "h (m)"],
    # Marine has a SINGLE combined column "DIMENSION (mm) L x W x H".
    # We capture it under a virtual field then split during row parsing.
    "_dimension_lxwxh_mm": [
        "dimension (mm) l x w x h",
        "dimension (mm)",
        "dimensions (mm)",
        "dimension mm",
        "l x w x h",
    ],
    "dry_weight_mt": [
        "dry weight per unit", "dry wt",
        "weight (mt) (dry)", "weight(mt) (dry)", "weight (dry)", "dry weight",
    ],
    "operating_weight_mt": [
        "operating weight per unit", "ope wt", "op weight",
        "weight (mt) (oper)", "weight(mt) (oper)", "weight (oper)",
        "operating weight", "operational weight",
        # Bare fallback for a single unqualified "WEIGHT (MT)" column (no
        # dry/oper split). Safe as a LAST resort only because dry_weight_mt
        # is checked earlier in this dict and has its own "(dry)"-qualified
        # patterns — those always win first on a document that qualifies
        # both columns, so this only fires when there's just one column.
        "weight (mt)",
    ],
    "hydrotest_weight_mt": ["hydrotest weight"],
    "pid": ["p&id", "pfd number", "p & id"],
    "remarks": ["remarks", "particulars"],   # Marine uses "PARTICULARS" for engineering notes
    "total_dry_weight_mt": ["total dry wt", "total dry weight"],
    "total_operating_weight_mt": ["total ope wt", "total operating weight"],

    # Extra fields from vendor drawings — added so a re-imported Excel
    # (whether it came from our own export or a contractor-authored
    # workbook that includes these extended columns) preserves the
    # values instead of dropping them.
    "length_overall_m": [
        "overall length", "l overall", "overall l",
        "length overall", "l (overall)",
    ],
    "mdmt_c": [
        "mdmt", "min design metal temp", "minimum design metal temp",
        "min design metal temperature", "minimum design metal temperature",
    ],
    "hydrostatic_test_press_barg": [
        "hydro test press", "hydrostatic test press", "hydrotest press",
        "hydrostatic test pressure", "hydrotest pressure",
    ],
    "insulation": [
        "insulation", "insulation type", "insulation & thickness",
        "insulation thickness",
    ],

    # Marine MEL lifecycle dropdowns — three adjacent columns. Each
    # equipment row marks ZERO or ONE (occasionally more) with a Y / X /
    # ✓ / "1" / the literal label. We capture each into its own virtual
    # field then collapse into a single `lifecycle_status` string below
    # so the model stays a single column instead of three booleans.
    "_lc_scrapped":     ["scrapped"],
    "_lc_refurbished":  ["refurbished", "refurbish", "refurb"],
    "_lc_new":          ["new"],   # plain "new" is rare elsewhere — safe here
}


# A cell in a SCRAPPED / REFURBISHED / NEW column counts as "marked"
# when it has ANY non-blank, non-zero, non-"N" value. Typical markers
# Marine-MEL spreadsheets use: "Y", "X", "✓", "✔", "1", "TRUE", or the
# label itself echoed back ("NEW", "SCRAPPED"). Treat anything else
# (and explicit "N", "0", "FALSE", "-") as unmarked.
_LC_NEGATIVE = {"", "-", "n", "no", "0", "false"}


def _is_lifecycle_marked(value: str | None) -> bool:
    if value is None:
        return False
    v = str(value).strip().lower()
    return v not in _LC_NEGATIVE


# Matches "7000 x 3000 x 3700" — three numbers separated by 'x' (or × or *).
# Captures all three so we can split into length / width / height.
_DIM_LXWXH_RE = re.compile(
    r"(\d+(?:\.\d+)?)\s*[x×*]\s*(\d+(?:\.\d+)?)\s*[x×*]\s*(\d+(?:\.\d+)?)",
    re.IGNORECASE,
)


def _split_lxwxh_mm(value: str) -> dict[str, str]:
    """If ``value`` looks like ``"7000 x 3000 x 3700"`` (or with units),
    split into the three MEL dimension columns and convert mm → m.
    Otherwise returns an empty dict (caller leaves the field alone).
    """
    if not value:
        return {}
    m = _DIM_LXWXH_RE.search(value)
    if not m:
        return {}
    try:
        l_mm, w_mm, h_mm = (float(m.group(i)) for i in (1, 2, 3))
    except ValueError:
        return {}
    # Convert mm → m, keep up to 3 decimals.
    def mm_to_m_str(x: float) -> str:
        return f"{(x / 1000):g}"
    return {
        "length_m":   mm_to_m_str(l_mm),
        "width_id_m": mm_to_m_str(w_mm),
        "height_tt_m": mm_to_m_str(h_mm),
    }

TAG_RE = re.compile(r"^[A-Z]{1,3}-[A-Z0-9]{2,}", re.IGNORECASE)


def _normalize_header(v: Any) -> str:
    return re.sub(r"\s+", " ", str(v or "")).strip().lower()


def _find_header_row(ws) -> int | None:
    """Locate the row that has the most matches against FIELD_HEADER_PATTERNS.

    Threshold is 4 — Marine MELs have fewer Topside-style columns so they
    score lower; 4 is enough to confidently identify a header without
    matching a single-stray-keyword row.
    """
    best_row, best_score = None, 0
    max_search = min(ws.max_row, 20)
    for r in range(1, max_search + 1):
        score = 0
        for cell in ws[r]:
            h = _normalize_header(cell.value)
            if not h:
                continue
            for patterns in FIELD_HEADER_PATTERNS.values():
                if any(p in h for p in patterns):
                    score += 1
                    break
        if score > best_score:
            best_score = score
            best_row = r
    return best_row if best_score >= 4 else None


def _build_column_map(ws, header_row: int) -> dict[str, int]:
    col_for_field: dict[str, int] = {}
    for cell in ws[header_row]:
        h = _normalize_header(cell.value)
        if not h:
            continue
        for field, patterns in FIELD_HEADER_PATTERNS.items():
            if field in col_for_field:
                continue
            if any(p in h for p in patterns):
                col_for_field[field] = cell.column
                break
    return col_for_field


def _stringify(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _looks_like_data_row(row_values: dict[str, Any]) -> bool:
    tag = row_values.get("client_tag")
    if not tag or not isinstance(tag, str):
        return False
    return bool(TAG_RE.match(tag.strip()))


def extract_equipment_rows(path: str, sheet_name: str | None = None) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    target_sheet = sheet_name or _pick_sheet(wb)
    ws = wb[target_sheet]

    header_row = _find_header_row(ws)
    if header_row is None:
        return []

    col_map = _build_column_map(ws, header_row)
    if "client_tag" not in col_map:
        # Some sheets put the tag in "old_tag"-style column; fall back.
        if "old_tag" in col_map:
            col_map["client_tag"] = col_map["old_tag"]
        else:
            return []

    out: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        row_data: dict[str, Any] = {}
        for field, col in col_map.items():
            row_data[field] = _stringify(ws.cell(row=r, column=col).value)
        if not _looks_like_data_row(row_data):
            continue
        row_data["client_tag"] = row_data["client_tag"].strip()

        # Marine sheets carry dimensions in a single "L x W x H (mm)" cell.
        # If we captured one of those, split it into the three real columns
        # AND drop the virtual field before insert. Existing L/W/H columns
        # from a Topsides sheet win — we only fill from the combined cell
        # when they're not already populated.
        combo = row_data.pop("_dimension_lxwxh_mm", None)
        if combo:
            split = _split_lxwxh_mm(combo)
            for k, v in split.items():
                if not row_data.get(k):
                    row_data[k] = v

        # Collapse the three Marine lifecycle columns into a single
        # `lifecycle_status` value. Whichever box(es) the row marks are
        # joined with " / ". Drop the virtual fields so they don't reach
        # the model.
        lc_scrapped = row_data.pop("_lc_scrapped", None)
        lc_refurbished = row_data.pop("_lc_refurbished", None)
        lc_new = row_data.pop("_lc_new", None)
        lc_flags: list[str] = []
        if _is_lifecycle_marked(lc_scrapped):
            lc_flags.append("SCRAPPED")
        if _is_lifecycle_marked(lc_refurbished):
            lc_flags.append("REFURBISHED")
        if _is_lifecycle_marked(lc_new):
            lc_flags.append("NEW")
        if lc_flags:
            row_data["lifecycle_status"] = " / ".join(lc_flags)

        # Backfill TOTAL weights from per-unit × count(configuration).
        # The reference EPC template stores DRY / OPE WT as per-unit and
        # TOTAL DRY / TOTAL OPE WT as the installed weight. Many source
        # workbooks leave the TOTAL columns blank when the config is
        # "1 x 100%" (per-unit = total anyway) or leave them blank as
        # a data-entry shortcut expecting the reader to compute. We
        # backfill them here — but ONLY when the source cell is empty.
        # If the source has an explicit TOTAL (even one that disagrees
        # with per-unit × count, which sometimes happens for standby
        # units), that value wins.
        from app.services.quantity import compute_installed_weight, is_blank
        cfg = row_data.get("configuration")
        if is_blank(row_data.get("total_dry_weight_mt")):
            computed = compute_installed_weight(row_data.get("dry_weight_mt"), cfg)
            if computed is not None:
                row_data["total_dry_weight_mt"] = computed
        if is_blank(row_data.get("total_operating_weight_mt")):
            computed = compute_installed_weight(row_data.get("operating_weight_mt"), cfg)
            if computed is not None:
                row_data["total_operating_weight_mt"] = computed

        # Preserve full row content in `data.raw` for forward compatibility
        full = {
            f"col_{c}": _stringify(ws.cell(row=r, column=c).value)
            for c in range(1, ws.max_column + 1)
        }
        row_data["__raw"] = full
        out.append(row_data)
    return out


def _pick_sheet(wb) -> str:
    for name in wb.sheetnames:
        if name.strip().lower() in ("equipment list", "topside equipment list"):
            return name
    # else: the sheet with the largest used range
    return max(wb.sheetnames, key=lambda n: wb[n].max_row * wb[n].max_column)
