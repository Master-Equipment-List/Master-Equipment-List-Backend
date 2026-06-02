"""Excel parser — reads every sheet into a list-of-rows JSON structure."""
from pathlib import Path

import openpyxl

from app.parsers.base import ParseResult


def parse_excel(path: str | Path) -> ParseResult:
    res = ParseResult(parser="excel")
    try:
        wb = openpyxl.load_workbook(filename=str(path), data_only=True, read_only=True)
    except Exception as e:
        res.status = "error"
        res.error = f"openpyxl failed: {e}"
        return res

    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows: list[list] = []
        for row in ws.iter_rows(values_only=True):
            rows.append([_normalize(c) for c in row])
        sheets[name] = {
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "rows": rows,
        }
    wb.close()
    res.data = {"sheets": sheets, "sheet_names": list(sheets.keys())}
    return res


def _normalize(v):
    if v is None:
        return None
    if isinstance(v, (int, float, bool, str)):
        return v
    return str(v)
